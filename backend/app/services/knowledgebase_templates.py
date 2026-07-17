from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from uuid import UUID

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import (
    Activity,
    ActivitySynonym,
    BOQItem,
    Project,
    ProjectKnowledgeUpload,
    ProjectLocation,
    ProjectScheduleItem,
    Unit,
    User,
)

SUPPORTED_TEMPLATE_TYPES = {
    "units",
    "activities",
    "locations",
    "boq",
    "schedule",
}


TEMPLATE_DEFINITIONS: dict[str, dict[str, list[str]]] = {
    "units": {
        "headers": ["name", "symbol", "unit_type"],
        "sample": ["Square Meter", "sqm", "area"],
    },
    "activities": {
        "headers": ["name", "category", "default_unit_symbol", "synonyms"],
        "sample": ["Plastering", "Finishing", "sqm", "plaster|plaster complete hua"],
    },
    "locations": {
        "headers": ["name", "location_type", "parent_name", "code"],
        "sample": ["Tower A", "tower", "", "T-A"],
    },
    "boq": {
        "headers": [
            "item_code",
            "item_description",
            "planned_quantity",
            "unit_symbol",
            "material_name",
            "activity_name",
        ],
        "sample": ["PL-001", "Internal plastering", "500", "sqm", "Cement", "Plastering"],
    },
    "schedule": {
        "headers": [
            "activity_name",
            "planned_start_date",
            "planned_end_date",
            "area_name",
            "sub_area_name",
            "planned_quantity",
            "unit_symbol",
        ],
        "sample": ["Plastering", "2026-07-20", "2026-07-25", "Tower A", "Floor 2", "100", "sqm"],
    },
}


@dataclass
class ImportSummary:
    upload_id: UUID
    upload_type: str
    status: str
    imported_count: int
    skipped_count: int
    error_count: int
    errors: list[str]


def build_template_workbook(template_type: str) -> bytes:
    template_type = _normalize_template_type(template_type)
    definition = TEMPLATE_DEFINITIONS[template_type]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = template_type
    sheet.append(definition["headers"])
    sheet.append(definition["sample"])

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = max(max_length + 2, 14)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.read()


def import_template(
    db: Session,
    company_id: UUID,
    project_id: UUID,
    upload_type: str,
    file_name: str,
    content: bytes,
    uploaded_by: UUID | None = None,
) -> ImportSummary:
    upload_type = _normalize_template_type(upload_type)
    _require_project(db, company_id, project_id)
    if uploaded_by:
        _require_user(db, company_id, uploaded_by)

    upload = ProjectKnowledgeUpload(
        company_id=company_id,
        project_id=project_id,
        uploaded_by=uploaded_by,
        upload_type=upload_type,
        file_name=file_name,
        status="processing",
    )
    db.add(upload)
    db.flush()

    rows, errors = _read_rows(content, upload_type)
    if errors:
        upload.status = "failed"
        upload.error_summary = "\n".join(errors[:50])
        db.commit()
        db.refresh(upload)
        return ImportSummary(
            upload_id=upload.id,
            upload_type=upload_type,
            status=upload.status,
            imported_count=0,
            skipped_count=0,
            error_count=len(errors),
            errors=errors,
        )

    try:
        imported_count, skipped_count, row_errors = _import_rows(
            db=db,
            company_id=company_id,
            project_id=project_id,
            upload_type=upload_type,
            rows=rows,
        )
    except ValueError as exc:
        row_errors = [str(exc)]
        imported_count = 0
        skipped_count = 0

    if row_errors:
        upload.status = "failed"
        upload.error_summary = "\n".join(row_errors[:50])
        db.rollback()
        failed_upload = ProjectKnowledgeUpload(
            company_id=company_id,
            project_id=project_id,
            uploaded_by=uploaded_by,
            upload_type=upload_type,
            file_name=file_name,
            status="failed",
            error_summary="\n".join(row_errors[:50]),
        )
        db.add(failed_upload)
        db.commit()
        db.refresh(failed_upload)
        return ImportSummary(
            upload_id=failed_upload.id,
            upload_type=upload_type,
            status="failed",
            imported_count=0,
            skipped_count=skipped_count,
            error_count=len(row_errors),
            errors=row_errors,
        )

    upload.status = "imported"
    db.commit()
    db.refresh(upload)
    return ImportSummary(
        upload_id=upload.id,
        upload_type=upload_type,
        status=upload.status,
        imported_count=imported_count,
        skipped_count=skipped_count,
        error_count=0,
        errors=[],
    )


def _normalize_template_type(template_type: str) -> str:
    normalized = template_type.strip().lower().replace("-", "_")
    aliases = {
        "boq_items": "boq",
        "project_schedule": "schedule",
        "schedule_items": "schedule",
        "project_locations": "locations",
    }
    normalized = aliases.get(normalized, normalized)
    if normalized not in SUPPORTED_TEMPLATE_TYPES:
        supported = ", ".join(sorted(SUPPORTED_TEMPLATE_TYPES))
        raise ValueError(f"Unsupported template type '{template_type}'. Supported: {supported}.")
    return normalized


def _read_rows(content: bytes, upload_type: str) -> tuple[list[dict[str, object]], list[str]]:
    errors: list[str] = []
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:  # noqa: BLE001
        return [], [f"Could not read Excel file: {exc}"]

    sheet = workbook.active
    header_values = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(max_row=1))]
    headers = [_normalize_header(value) for value in header_values]
    required_headers = TEMPLATE_DEFINITIONS[upload_type]["headers"]
    missing_headers = [header for header in required_headers if header not in headers]
    if missing_headers:
        errors.append(f"Missing required columns: {', '.join(missing_headers)}")
        return [], errors

    rows: list[dict[str, object]] = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row_data = {
            headers[index]: value
            for index, value in enumerate(row[: len(headers)])
            if headers[index]
        }
        if all(_is_blank(value) for value in row_data.values()):
            continue
        row_data["_row_number"] = row_number
        rows.append(row_data)

    if not rows:
        errors.append("Template has no data rows.")

    return rows, errors


def _import_rows(
    db: Session,
    company_id: UUID,
    project_id: UUID,
    upload_type: str,
    rows: list[dict[str, object]],
) -> tuple[int, int, list[str]]:
    if upload_type == "units":
        return _import_units(db, company_id, rows)
    if upload_type == "activities":
        return _import_activities(db, company_id, rows)
    if upload_type == "locations":
        return _import_locations(db, company_id, project_id, rows)
    if upload_type == "boq":
        return _import_boq(db, company_id, project_id, rows)
    if upload_type == "schedule":
        return _import_schedule(db, company_id, project_id, rows)
    raise ValueError(f"Unsupported upload type: {upload_type}")


def _import_units(
    db: Session,
    company_id: UUID,
    rows: list[dict[str, object]],
) -> tuple[int, int, list[str]]:
    imported = 0
    skipped = 0
    errors: list[str] = []
    for row in rows:
        row_number = row["_row_number"]
        name = _text(row.get("name"))
        symbol = _text(row.get("symbol"))
        if not name or not symbol:
            errors.append(f"Row {row_number}: name and symbol are required.")
            continue
        if _find_unit_by_symbol(db, company_id, symbol):
            skipped += 1
            continue
        db.add(Unit(company_id=company_id, name=name, symbol=symbol, unit_type=_text(row.get("unit_type"))))
        imported += 1
    return imported, skipped, errors


def _import_activities(
    db: Session,
    company_id: UUID,
    rows: list[dict[str, object]],
) -> tuple[int, int, list[str]]:
    imported = 0
    skipped = 0
    errors: list[str] = []
    for row in rows:
        row_number = row["_row_number"]
        name = _text(row.get("name"))
        if not name:
            errors.append(f"Row {row_number}: activity name is required.")
            continue
        unit = None
        unit_symbol = _text(row.get("default_unit_symbol"))
        if unit_symbol:
            unit = _find_unit_by_symbol(db, company_id, unit_symbol)
            if not unit:
                errors.append(f"Row {row_number}: unit symbol '{unit_symbol}' was not found.")
                continue
        activity = _find_activity_by_name(db, company_id, name)
        if activity:
            skipped += 1
        else:
            activity = Activity(
                company_id=company_id,
                name=name,
                category=_text(row.get("category")),
                default_unit_id=unit.id if unit else None,
            )
            db.add(activity)
            db.flush()
            imported += 1
        synonyms = _split_synonyms(row.get("synonyms"))
        for synonym in synonyms:
            if not _find_synonym(db, activity.id, synonym):
                db.add(
                    ActivitySynonym(
                        company_id=company_id,
                        activity_id=activity.id,
                        synonym=synonym,
                        language=None,
                    )
                )
    return imported, skipped, errors


def _import_locations(
    db: Session,
    company_id: UUID,
    project_id: UUID,
    rows: list[dict[str, object]],
) -> tuple[int, int, list[str]]:
    imported = 0
    skipped = 0
    errors: list[str] = []
    for row in rows:
        row_number = row["_row_number"]
        name = _text(row.get("name"))
        location_type = _text(row.get("location_type"))
        parent_name = _text(row.get("parent_name"))
        if not name or not location_type:
            errors.append(f"Row {row_number}: name and location_type are required.")
            continue
        parent = None
        if parent_name:
            parent = _find_location_by_name(db, company_id, project_id, parent_name)
            if not parent:
                errors.append(f"Row {row_number}: parent location '{parent_name}' was not found.")
                continue
        if _find_location_by_name(db, company_id, project_id, name, parent.id if parent else None):
            skipped += 1
            continue
        db.add(
            ProjectLocation(
                company_id=company_id,
                project_id=project_id,
                parent_location_id=parent.id if parent else None,
                name=name,
                location_type=location_type,
                code=_text(row.get("code")),
            )
        )
        db.flush()
        imported += 1
    return imported, skipped, errors


def _import_boq(
    db: Session,
    company_id: UUID,
    project_id: UUID,
    rows: list[dict[str, object]],
) -> tuple[int, int, list[str]]:
    imported = 0
    skipped = 0
    errors: list[str] = []
    for row in rows:
        row_number = row["_row_number"]
        description = _text(row.get("item_description"))
        if not description:
            errors.append(f"Row {row_number}: item_description is required.")
            continue
        unit = _optional_unit(db, company_id, row, "unit_symbol", errors)
        activity = _optional_activity(db, company_id, row, "activity_name", errors)
        if errors and errors[-1].startswith(f"Row {row_number}:"):
            continue
        item_code = _text(row.get("item_code"))
        if item_code and _find_boq_by_code(db, project_id, item_code):
            skipped += 1
            continue
        db.add(
            BOQItem(
                company_id=company_id,
                project_id=project_id,
                item_code=item_code,
                item_description=description,
                planned_quantity=_decimal(row.get("planned_quantity")),
                unit_id=unit.id if unit else None,
                material_name=_text(row.get("material_name")),
                activity_id=activity.id if activity else None,
            )
        )
        imported += 1
    return imported, skipped, errors


def _import_schedule(
    db: Session,
    company_id: UUID,
    project_id: UUID,
    rows: list[dict[str, object]],
) -> tuple[int, int, list[str]]:
    imported = 0
    skipped = 0
    errors: list[str] = []
    for row in rows:
        row_number = row["_row_number"]
        activity_name = _text(row.get("activity_name"))
        if not activity_name:
            errors.append(f"Row {row_number}: activity_name is required.")
            continue
        activity = _find_activity_by_name(db, company_id, activity_name)
        unit = _optional_unit(db, company_id, row, "unit_symbol", errors)
        area = _optional_location(db, company_id, project_id, row, "area_name", errors)
        sub_area = _optional_location(db, company_id, project_id, row, "sub_area_name", errors)
        if errors and errors[-1].startswith(f"Row {row_number}:"):
            continue
        db.add(
            ProjectScheduleItem(
                company_id=company_id,
                project_id=project_id,
                activity_id=activity.id if activity else None,
                activity_name=activity_name,
                planned_start_date=_date(row.get("planned_start_date")),
                planned_end_date=_date(row.get("planned_end_date")),
                area_id=area.id if area else None,
                sub_area_id=sub_area.id if sub_area else None,
                planned_quantity=_decimal(row.get("planned_quantity")),
                unit_id=unit.id if unit else None,
            )
        )
        imported += 1
    return imported, skipped, errors


def _optional_unit(
    db: Session,
    company_id: UUID,
    row: dict[str, object],
    key: str,
    errors: list[str],
) -> Unit | None:
    symbol = _text(row.get(key))
    if not symbol:
        return None
    unit = _find_unit_by_symbol(db, company_id, symbol)
    if not unit:
        errors.append(f"Row {row['_row_number']}: unit symbol '{symbol}' was not found.")
    return unit


def _optional_activity(
    db: Session,
    company_id: UUID,
    row: dict[str, object],
    key: str,
    errors: list[str],
) -> Activity | None:
    name = _text(row.get(key))
    if not name:
        return None
    activity = _find_activity_by_name(db, company_id, name)
    if not activity:
        errors.append(f"Row {row['_row_number']}: activity '{name}' was not found.")
    return activity


def _optional_location(
    db: Session,
    company_id: UUID,
    project_id: UUID,
    row: dict[str, object],
    key: str,
    errors: list[str],
) -> ProjectLocation | None:
    name = _text(row.get(key))
    if not name:
        return None
    location = _find_location_by_name(db, company_id, project_id, name)
    if not location:
        errors.append(f"Row {row['_row_number']}: location '{name}' was not found.")
    return location


def _require_project(db: Session, company_id: UUID, project_id: UUID) -> Project:
    project = db.get(Project, project_id)
    if not project or project.company_id != company_id:
        raise ValueError("Project not found for this company.")
    return project


def _require_user(db: Session, company_id: UUID, user_id: UUID) -> User:
    user = db.get(User, user_id)
    if not user or user.company_id != company_id:
        raise ValueError("User not found for this company.")
    return user


def _find_unit_by_symbol(db: Session, company_id: UUID, symbol: str) -> Unit | None:
    return db.scalar(
        select(Unit)
        .where(Unit.company_id == company_id)
        .where(Unit.symbol.ilike(symbol.strip()))
    )


def _find_activity_by_name(db: Session, company_id: UUID, name: str) -> Activity | None:
    return db.scalar(
        select(Activity)
        .where(Activity.company_id == company_id)
        .where(Activity.name.ilike(name.strip()))
    )


def _find_synonym(db: Session, activity_id: UUID, synonym: str) -> ActivitySynonym | None:
    return db.scalar(
        select(ActivitySynonym)
        .where(ActivitySynonym.activity_id == activity_id)
        .where(ActivitySynonym.synonym.ilike(synonym.strip()))
    )


def _find_location_by_name(
    db: Session,
    company_id: UUID,
    project_id: UUID,
    name: str,
    parent_location_id: UUID | None = None,
) -> ProjectLocation | None:
    statement = (
        select(ProjectLocation)
        .where(ProjectLocation.company_id == company_id)
        .where(ProjectLocation.project_id == project_id)
        .where(ProjectLocation.name.ilike(name.strip()))
    )
    if parent_location_id:
        statement = statement.where(ProjectLocation.parent_location_id == parent_location_id)
    return db.scalar(statement)


def _find_boq_by_code(db: Session, project_id: UUID, item_code: str) -> BOQItem | None:
    return db.scalar(
        select(BOQItem)
        .where(BOQItem.project_id == project_id)
        .where(BOQItem.item_code == item_code)
    )


def _normalize_header(value: str) -> str:
    return value.strip().lower().replace(" ", "_").replace("-", "_")


def _text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_blank(value: object) -> bool:
    return value is None or str(value).strip() == ""


def _split_synonyms(value: object) -> list[str]:
    text = _text(value)
    if not text:
        return []
    return [part.strip() for part in text.split("|") if part.strip()]


def _decimal(value: object) -> Decimal | None:
    if _is_blank(value):
        return None
    try:
        return Decimal(str(value))
    except InvalidOperation:
        return None


def _date(value: object) -> date | None:
    if _is_blank(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value).strip())
